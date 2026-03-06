"""資料處理函式模組"""
import pandas as pd
import numpy as np
import toad
from typing import List, Dict, Tuple
import tqdm
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except:
    HAS_OPENPYXL = False


def get_dataset(data_pth: str, date_colName: str, y_colName: str, 
                org_colName: str, data_encode: str, key_colNames: List[str],
                drop_colNames: List[str] = None,
                miss_vals: List[int] = None) -> pd.DataFrame:
    """載入並格式化資料
    
    參數:
        data_pth: 資料檔案路徑
        date_colName: 日期欄位名稱
        y_colName: 標籤欄位名稱
        org_colName: 機構欄位名稱
        data_encode: 資料編碼
        key_colNames: 主鍵欄位 (用於去重)
        drop_colNames: 需剔除的欄位
        miss_vals: 需替換為 NaN 的異常值清單，預設為 [-1, -999, -1111]
    """
    if drop_colNames is None:
        drop_colNames = []
    if miss_vals is None:
        miss_vals = [-1, -999, -1111]
    
    # 支援多種格式讀取
    for fmt, reader in [('parquet', pd.read_parquet), ('csv', pd.read_csv), 
                         ('xlsx', pd.read_excel), ('pkl', pd.read_pickle)]:
        try:
            data = reader(data_pth)
            break
        except:
            continue
    
    # 將異常值替換為 NaN
    data.replace({v: np.nan for v in miss_vals}, inplace=True)
    
    # 去重與過濾
    data = data[data[y_colName].isin([0, 1])]
    data = data.drop_duplicates(subset=key_colNames)
    
    # 剔除無效欄位
    data.drop(columns=[c for c in drop_colNames if c in data.columns], errors='ignore')
    data.drop(columns=[c for c in data.columns if data[c].nunique() <= 1], errors='ignore')
    
    # 重命名欄位
    data.rename(columns={date_colName: 'new_date', y_colName: 'new_target', 
                         org_colName: 'new_org'}, inplace=True)
    data['new_date'] = data['new_date'].astype(str).str.replace('-', '', regex=False).str[:8]
    data['new_date_ym'] = data['new_date'].str[:6]
    
    return data


def org_analysis(data: pd.DataFrame, oos_orgs: List[str] = None) -> pd.DataFrame:
    """機構樣本統計分析
    
    參數:
        data: 資料
        oos_orgs: 貸外機構清單，用於識別 OOS 樣本
    """
    stat = data.groupby(['new_org', 'new_date_ym']).agg(
        單月壞樣本數=('new_target', 'sum'),
        單月總樣本數=('new_target', 'count'),
        單月壞樣率=('new_target', 'mean')
    ).reset_index()
    
    # 累計統計
    stat['總壞樣本數'] = stat.groupby('new_org')['單月壞樣本數'].transform('sum')
    stat['總樣本數'] = stat.groupby('new_org')['單月總樣本數'].transform('sum')
    stat['總壞樣率'] = stat['總壞樣本數'] / stat['總樣本數']
    
    # 標記是否為 OOS 機構
    if oos_orgs and len(oos_orgs) > 0:
        stat['樣本類型'] = stat['new_org'].apply(lambda x: '貸外' if x in oos_orgs else '建模')
    else:
        stat['樣本類型'] = '建模'
    
    stat = stat.rename(columns={'new_org': '機構', 'new_date_ym': '年月'})
    
    # 按樣本類型排序 (建模優先，貸外最後)
    stat = stat.sort_values(['樣本類型', '機構', '年月'], ascending=[True, True, True])
    stat = stat.reset_index(drop=True)
    
    return stat[['機構', '年月', '單月壞樣本數', '單月總樣本數', '單月壞樣率', '總壞樣本數', '總樣本數', '總壞樣率', '樣本類型']]


def missing_check(data: pd.DataFrame, channel: Dict[str, List[str]] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """計算缺失率 - 包含整體及分機構缺失率
    
    返回:
        miss_detail: 缺失率明細 (格式：變數、整體、org1, org2, ..., orgn)
        miss_ch: 整體缺失率 (每個變數的整體缺失率)
    """
    miss_vals = [-1, -999, -1111]
    miss_ch = []
    
    # 排除非變數欄位：record_id, target, org_info 等
    exclude_cols = ['new_date', 'new_date_ym', 'new_target', 'new_org', 'record_id', 'target', 'org_info']
    cols = [c for c in data.columns if c not in exclude_cols]
    
    # 計算整體缺失率
    for col in tqdm.tqdm(cols, desc="缺失率"):
        rate = ((data[col].isin(miss_vals)) | (data[col].isna())).mean()
        miss_ch.append({'變數': col, '整體缺失率': round(rate, 4)})
    
    miss_ch = pd.DataFrame(miss_ch)
    
    # 計算分機構缺失率並轉換為寬格式
    orgs = sorted(data['new_org'].unique())
    miss_detail_dict = {'變數': []}
    miss_detail_dict['整體'] = []
    
    for org in orgs:
        miss_detail_dict[org] = []
    
    for col in cols:
        miss_detail_dict['變數'].append(col)
        # 整體缺失率
        overall_rate = ((data[col].isin(miss_vals)) | (data[col].isna())).mean()
        miss_detail_dict['整體'].append(round(overall_rate, 4))
        
        # 各機構缺失率
        for org in orgs:
            org_data = data[data['new_org'] == org]
            rate = ((org_data[col].isin(miss_vals)) | (org_data[col].isna())).mean()
            miss_detail_dict[org].append(round(rate, 4))
    
    miss_detail = pd.DataFrame(miss_detail_dict)
    # 按整體缺失率降序排序
    miss_detail = miss_detail.sort_values('整體', ascending=False)
    miss_detail = miss_detail.reset_index(drop=True)
    
    return miss_detail, miss_ch


def calculate_iv(data: pd.DataFrame, features: List[str], n_jobs: int = 4) -> pd.DataFrame:
    """計算 IV 值 - 使用 toad.transform.Combiner 進行分箱，分箱數設為 5，保留 NaN 值"""
    import tqdm
    from joblib import Parallel, delayed
    
    def _calc_iv(f):
        try:
            # 使用 toad.transform.Combiner 進行分箱，分箱數設為 5
            c = toad.transform.Combiner()
            data_temp = data[[f, 'new_target']].copy()
            data_temp.columns = ['x', 'y']
            data_temp['x_bin'] = c.fit_transform(X=data_temp['x'], y=data_temp['y'], method='dt', n_bins=5, min_samples=0.05/5, empty_separate=True)
            
            # 使用分箱後的資料計算 IV 值
            iv_df = toad.quality(data_temp[['x_bin', 'y']], 'y', iv_only=True)
            if 'iv' in iv_df.columns and len(iv_df) > 0:
                iv_value = iv_df['iv'].iloc[0]
                if not np.isnan(iv_value):
                    return {'變數': f, 'IV': round(iv_value, 4)}
            return None
        except Exception as e:
            print(f"   IV 計算錯誤：變數={f}, 錯誤={e}")
            return None
    
    # 使用 tqdm 顯示進度
    results = Parallel(n_jobs=n_jobs, verbose=0)(
        delayed(_calc_iv)(f) for f in features
    )
    iv_list = [r for r in results if r is not None]
    
    if len(iv_list) == 0:
        print(f"   IV 計算結果為空，特徵數量={len(features)}")
        return pd.DataFrame(columns=['變數', 'IV'])
    
    return pd.DataFrame(iv_list).sort_values('IV', ascending=False)


def calculate_corr(data: pd.DataFrame, features: List[str]) -> pd.DataFrame:
    """計算相關性矩陣"""
    corr = data[features].corr().abs()
    return corr


def export_report_xlsx(filepath: str, data_name: str, data: pd.DataFrame, 
                       sheet_name: str, description: str = ""):
    """匯出 xlsx 報告 - 支援追加"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.create_sheet(sheet_name)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    
    # 寫入描述
    ws['A1'] = f"資料：{data_name}"
    ws['A2'] = f"時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    if description:
        ws['A3'] = f"描述：{description}"
    
    # 寫入資料
    start_row = 5
    for i, col in enumerate(data.columns):
        ws.cell(start_row, i+1, col)
    
    for i, row in enumerate(data.values):
        for j, val in enumerate(row):
            ws.cell(start_row+1+i, j+1, val)
    
    # 樣式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    wb.save(filepath)
    print(f"[{sheet_name}] 已儲存至 {filepath}")
