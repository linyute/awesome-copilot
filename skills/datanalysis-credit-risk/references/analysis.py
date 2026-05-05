"""變數選擇與分析模組 - 簡化版本
PSI 計算在 func.py 中重複使用，analysis.py 僅處理變數選擇
"""
import pandas as pd
import numpy as np
import toad
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_auc_score
from joblib import Parallel, delayed


def drop_abnormal_ym(data: pd.DataFrame, min_ym_bad_sample: int = 1,
                     min_ym_sample: int = 500) -> tuple:
    """篩選異常月份 - 全域統計，非按機構"""
    stat = data.groupby('new_date_ym').agg(
        bad_cnt=('new_target', 'sum'),
        total=('new_target', 'count')
    ).reset_index()

    abnormal = stat[(stat['bad_cnt'] < min_ym_bad_sample) | (stat['total'] < min_ym_sample)]
    abnormal = abnormal.rename(columns={'new_date_ym': '年月'})
    abnormal['去除條件'] = abnormal.apply(
        lambda x: f'壞樣本數量 {x["bad_cnt"]} 小於 {min_ym_bad_sample}' if x['bad_cnt'] < min_ym_bad_sample else f'總樣本數量 {x["total"]} 小於 {min_ym_sample}', axis=1
    )

    if len(abnormal) > 0:
        data = data[~data['new_date_ym'].isin(abnormal['年月'])]

    # 移除空行
    abnormal = abnormal.dropna(how='all')
    abnormal = abnormal.reset_index(drop=True)

    return data, abnormal


def drop_highmiss_features(data: pd.DataFrame, miss_channel: pd.DataFrame,
                           threshold: float = 0.6) -> tuple:
    """剔除高缺失率特徵"""
    high_miss = miss_channel[miss_channel['整體缺失率'] > threshold].copy()
    high_miss['缺失率'] = high_miss['整體缺失率']

    # 修改移除條件以顯示具體的缺失率值
    high_miss['去除條件'] = high_miss.apply(
        lambda x: f'整體缺失率為 {x["缺失率"]:.4f}，超過閾值 {threshold}', axis=1
    )

    # 移除空行
    high_miss = high_miss.dropna(how='all')
    high_miss = high_miss.reset_index(drop=True)

    # 剔除高缺失率特徵
    if len(high_miss) > 0 and '變數' in high_miss.columns:
        to_drop = high_miss['變數'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns])

    return data, high_miss[['變數', '缺失率', '去除條件']]


def drop_lowiv_features(data: pd.DataFrame, features: List[str],
                       overall_iv_threshold: float = 0.05, org_iv_threshold: float = 0.02,
                       max_org_threshold: int = 8, n_jobs: int = 4) -> tuple:
    """剔除低 IV 特徵 - 多程序版本，返回 IV 明細和 IV 處理表

    參數:
        overall_iv_threshold: 整體 IV 閾值，低於此值的變數將記錄在 IV 處理表中
        org_iv_threshold: 單一機構 IV 閾值，低於此值則視為不滿足
        max_org_threshold: 最大容忍機構數量，若超過此數量的機構 IV 低於閾值，則記錄在 IV 處理表中

    返回:
        data: 剔除後的資料
        iv_detail: IV 明細 (每個特徵在每個機構及整體的 IV 值)
        iv_process: IV 處理表 (不符合條件的特徵)
    """
    from references.func import calculate_iv
    from joblib import Parallel, delayed

    orgs = sorted(data['new_org'].unique())

    print(f"   IV 計算：特徵數量={len(features)}，機構數量={len(orgs)}")

    # 一次計算所有機構的 IV 值
    def _calc_org_iv(org):
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, n_jobs=1)
        if len(org_iv) > 0:
            org_iv = org_iv.rename(columns={'IV': 'IV值'})
            org_iv['機構'] = org
            return org_iv
        return None

    # 計算整體 IV
    print(f"   正在計算整體 IV...")
    iv_overall = calculate_iv(data, features, n_jobs=n_jobs)
    print(f"   整體 IV 計算結果：{len(iv_overall)} 個特徵")
    if len(iv_overall) == 0:
        print(f"   警告：整體 IV 計算結果為空，返回空表")
        return data, pd.DataFrame(columns=['變數', 'IV值', '機構', '類型']), pd.DataFrame(columns=['變數', '整體IV', '低IV機構數', '處理原因'])
    iv_overall = iv_overall.rename(columns={'IV': 'IV值'})

    # 並行計算所有機構的 IV 值
    print(f"   正在並行計算 {len(orgs)} 個機構的 IV 值...")
    iv_by_org_results = Parallel(n_jobs=n_jobs, verbose=0)(
        delayed(_calc_org_iv)(org) for org in orgs
    )
    iv_by_org = [r for r in iv_by_org_results if r is not None]
    iv_by_org = pd.concat(iv_by_org, ignore_index=True) if iv_by_org else pd.DataFrame(columns=['變數', 'IV值', '機構'])
    print(f"   機構 IV 彙總：{len(iv_by_org)} 筆記錄")

    # 轉換為寬格式：特徵、整體、org1, org2, ..., orgn
    iv_detail_dict = {'變數': []}
    iv_detail_dict['整體'] = []

    for org in orgs:
        iv_detail_dict[org] = []

    # 獲取所有特徵
    all_vars = set(iv_overall['變數'].tolist())
    if len(iv_by_org) > 0:
        all_vars.update(iv_by_org['變數'].tolist())
    all_vars = sorted(all_vars)

    for var in all_vars:
        iv_detail_dict['變數'].append(var)

        # 整體 IV
        var_overall = iv_overall[iv_overall['變數'] == var]
        if len(var_overall) > 0:
            iv_detail_dict['整體'].append(var_overall['IV值'].values[0])
        else:
            iv_detail_dict['整體'].append(None)

        # 各機構 IV
        for org in orgs:
            var_org = iv_by_org[iv_by_org['機構'] == org]
            var_org = var_org[var_org['變數'] == var]
            if len(var_org) > 0:
                iv_detail_dict[org].append(var_org['IV值'].values[0])
            else:
                iv_detail_dict[org].append(None)

    iv_detail = pd.DataFrame(iv_detail_dict)
    # 按整體 IV 降序排序
    iv_detail = iv_detail.sort_values('整體', ascending=False)
    iv_detail = iv_detail.reset_index(drop=True)

    # 標記不符合條件的特徵
    # 1. 整體 IV 低於閾值
    iv_overall_low = iv_overall[iv_overall['IV值'] < overall_iv_threshold]['變數'].tolist()

    # 2. 單一機構 IV 低於閾值的機構數量
    if len(iv_by_org) > 0:
        iv_by_org_low = iv_by_org[iv_by_org['IV值'] < org_iv_threshold].groupby('變數').size().reset_index()
        iv_by_org_low.columns = ['變數', '低IV機構數']
    else:
        iv_by_org_low = pd.DataFrame(columns=['變數', '低IV機構數'])

    # 獲取每個特徵的低 IV 機構清單
    low_iv_orgs_dict = {}
    if len(iv_by_org) > 0:
        for var in iv_by_org['變數'].unique():
            var_orgs = iv_by_org[(iv_by_org['變數'] == var) & (iv_by_org['IV值'] < org_iv_threshold)]['機構'].tolist()
            low_iv_orgs_dict[var] = var_orgs

    # 3. 標記需要處理的特徵
    iv_process = []

    # 偵錯資訊：IV 分佈統計
    if len(iv_overall) > 0:
        print(f"   整體 IV 統計：最小值={iv_overall['IV值'].min():.4f}, 最大值={iv_overall['IV值'].max():.4f}, 中位數={iv_overall['IV值'].median():.4f}")
        print(f"   整體 IV 小於 {overall_iv_threshold} 的特徵數量：{(iv_overall['IV值'] < overall_iv_threshold).sum()}/{len(iv_overall)}")

    if len(iv_by_org_low) > 0:
        print(f"   機構 IV 小於 {org_iv_threshold} 的特徵統計：")
        print(f"     最大低 IV 機構數量：{iv_by_org_low['低IV機構數'].max()}")
        print(f"     低 IV 機構數量大於或等於 {max_org_threshold} 的特徵數量：{(iv_by_org_low['低IV機構數'] >= max_org_threshold).sum()}/{len(iv_by_org_low)}")

    for var in features:
        reasons = []

        # 檢查整體 IV
        var_overall_iv = iv_overall[iv_overall['變數'] == var]['IV值'].values
        if len(var_overall_iv) > 0 and var_overall_iv[0] < overall_iv_threshold:
            reasons.append(f'整體 IV {var_overall_iv[0]:.4f} 小於閾值 {overall_iv_threshold}')

        # 檢查機構 IV
        var_org_low = iv_by_org_low[iv_by_org_low['變數'] == var]
        if len(var_org_low) > 0 and var_org_low['低IV機構數'].values[0] >= max_org_threshold:
            reasons.append(f'在 {var_org_low["低IV機構數"].values[0]} 個機構中 IV 小於閾值 {org_iv_threshold}')

        if reasons:
            iv_process.append({
                '變數': var,
                '處理原因': '; '.join(reasons),
                '低IV機構': ','.join(low_iv_orgs_dict.get(var, []))
            })

    iv_process = pd.DataFrame(iv_process)
    iv_process = iv_process.reset_index(drop=True)

    # 剔除不符合條件的特徵
    if len(iv_process) > 0 and '變數' in iv_process.columns:
        to_drop = iv_process['變數'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns])

    return data, iv_detail, iv_process


def drop_highcorr_features(data: pd.DataFrame, features: List[str],
                           threshold: float = 0.8, gain_dict: dict = None, top_n_keep: int = 20) -> tuple:
    """剔除高相關性特徵 - 基於原始 gain，一次剔除一個特徵

    參數:
        data: 資料
        features: 特徵列表
        threshold: 相關性閾值
        gain_dict: 特徵到原始 gain 的映射字典
        top_n_keep: 按原始 gain 排名保留前 N 個特徵

    返回:
        data: 剔除後的資料
        dropped_info: 剔除資訊
    """
    if gain_dict is None:
        gain_dict = {}

    # 獲取當前特徵列表 (僅限資料中存在的特徵)
    current_features = [f for f in features if f in data.columns]

    if len(current_features) == 0:
        return data, pd.DataFrame(columns=['變數', '相關變數', '去除條件'])

    # 確定要保留的特徵 (原始 gain 前 N 名)
    if gain_dict:
        # 僅考慮當前剩餘特徵中的 gain
        current_gain_dict = {k: v for k, v in gain_dict.items() if k in current_features}
        if current_gain_dict:
            sorted_features = sorted(current_gain_dict.keys(), key=lambda x: current_gain_dict[x], reverse=True)
            top_features = set(sorted_features[:top_n_keep])
            # 建立特徵到排名的映射
            rank_dict = {v: i+1 for i, v in enumerate(sorted_features)}
        else:
            top_features = set()
            rank_dict = {}
    else:
        top_features = set()
        rank_dict = {}

    dropped_info = []

    # 迴圈剔除直到沒有高相關性特徵對
    while True:
        # 重新計算相關性矩陣 (僅針對當前剩餘特徵)
        current_features = [f for f in current_features if f in data.columns]
        if len(current_features) < 2:
            break

        corr = data[current_features].corr().abs()
        upper = corr.where(np.triu(np.ones(corr.shape), k=1).astype(bool))

        # 尋找所有高相關性特徵對
        high_corr_pairs = []
        for i, col1 in enumerate(upper.columns):
            for col2 in upper.columns[i+1:]:
                corr_val = upper.loc[col1, col2]
                if pd.notna(corr_val) and corr_val > threshold:
                    high_corr_pairs.append((col1, col2, corr_val))

        if not high_corr_pairs:
            break

        # 對於每一對高相關性特徵，選擇原始 gain 較小的特徵作為剔除候選
        candidates = set()
        for col1, col2, corr_val in high_corr_pairs:
            # 跳過前 N 個保留特徵
            if col1 in top_features and col2 in top_features:
                continue

            gain1 = gain_dict.get(col1, 0)
            gain2 = gain_dict.get(col2, 0)

            # 選擇原始 gain 較小的特徵
            if gain1 <= gain2:
                candidates.add(col1)
            else:
                candidates.add(col2)

        if not candidates:
            break

        # 從候選特徵中選擇原始 gain 最小的一個進行剔除
        candidates_list = list(candidates)
        candidates_with_gain = [(c, gain_dict.get(c, 0)) for c in candidates_list]
        candidates_with_gain.sort(key=lambda x: x[1])
        to_drop = candidates_with_gain[0][0]

        # 尋找與該特徵高度相關的所有特徵
        related_vars = []
        for col1, col2, corr_val in high_corr_pairs:
            if col1 == to_drop:
                related_vars.append((col2, corr_val))
            elif col2 == to_drop:
                related_vars.append((col1, corr_val))

        # 記錄剔除資訊
        # 相關變數欄位：顯示特徵名稱和相似度值 (相關性值)
        related_str = ','.join([f"{v}(相似度={c:.4f})" for v, c in related_vars])
        # 移除條件欄位：顯示相關特徵及其對應的 gain 值
        gain_str = ','.join([f"{v}(gain={gain_dict.get(v, 0):.2f})" for v, c in related_vars])
        dropped_info.append({
            '變數': to_drop,
            '原始gain': gain_dict.get(to_drop, 0),
            '原始gain排名': rank_dict.get(to_drop, '-'),
            '相關變數': related_str,
            '去除條件': gain_str
        })

        # 從資料中刪除該特徵
        data = data.drop(columns=[to_drop], errors='ignore')
        current_features.remove(to_drop)

        print(f"   已剔除特徵：{to_drop} (原始 gain={gain_dict.get(to_drop, 0):.2f})")

    # 轉換為 DataFrame 並按原始 gain 降序排序
    dropped_df = pd.DataFrame(dropped_info)
    if len(dropped_df) > 0:
        dropped_df = dropped_df.sort_values('原始gain', ascending=False)
        dropped_df = dropped_df.reset_index(drop=True)

    return data, dropped_df


def drop_highnoise_features(data: pd.DataFrame, features: List[str],
                           n_estimators: int = 100, max_depth: int = 5, gain_threshold: float = 50) -> tuple:
    """Null Importance 移除高雜訊特徵"""
    # 檢查特徵列表是否為空
    if len(features) == 0:
        print("   無特徵可處理")
        return data, pd.DataFrame(columns=['變數', '原始gain', '反轉後gain'])

    # 檢查資料是否充足
    if len(data) < 1000:
        print(f"   資料量不足 ({len(data)} 行)，跳過 Null Importance")
        return data, pd.DataFrame(columns=['變數', '原始gain', '反轉後gain'])

    X = data[features].copy()
    Y = data['new_target'].copy()

    # 檢查 X 是否為空或包含 NaN
    if X.shape[1] == 0:
        print("   特徵資料為空，跳過 Null Importance")
        return data, pd.DataFrame(columns=['變數', '原始gain', '反轉後gain'])

    # 填補 NaN
    X = X.fillna(0)

    # 打亂標籤
    Y_permuted = Y.copy()
    for _ in range(20):
        Y_permuted = np.random.permutation(Y_permuted)

    clf = lgb.LGBMClassifier(
        objective='binary', boosting_type='gbdt', learning_rate=0.05,
        max_depth=max_depth, min_child_samples=2000, min_child_weight=20,
        n_estimators=n_estimators, num_leaves=2**max_depth - 1, n_jobs=-1, verbose=-1
    )

    clf_permuted = lgb.LGBMClassifier(
        objective='binary', boosting_type='gbdt', learning_rate=0.05,
        max_depth=max_depth, min_child_samples=2000, min_child_weight=20,
        n_estimators=n_estimators, num_leaves=2**max_depth - 1, n_jobs=-1, verbose=-1
    )

    results, results_permuted = [], []

    print("正在進行 Null Importance 計算...")
    for i in range(2):
        random_n = np.random.randint(30)

        X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.3, random_state=random_n)

        # 檢查訓練資料是否有效
        if X_train.shape[0] == 0 or X_test.shape[0] == 0:
            print(f"  第 {i+1} 輪：資料切分失敗，跳過")
            continue

        clf.fit(X_train, y_train)

        X_train_, X_test_, y_train_, y_test_ = train_test_split(X, Y_permuted, test_size=0.3, random_state=random_n)

        if X_train_.shape[0] == 0 or X_test_.shape[0] == 0:
            print(f"  第 {i+1} 輪：打亂後資料切分失敗，跳過")
            continue

        clf_permuted.fit(X_train_, y_train_)

        imp_real = pd.DataFrame({
            'feature': clf.booster_.feature_name(),
            'gain': clf.booster_.feature_importance(importance_type='gain')
        })
        imp_perm = pd.DataFrame({
            'feature': clf_permuted.booster_.feature_name(),
            'gain': clf_permuted.booster_.feature_importance(importance_type='gain')
        })

        results.append(imp_real)
        results_permuted.append(imp_perm)

        train_auc = roc_auc_score(y_train, clf.predict_proba(X_train)[:, 1])
        test_auc = roc_auc_score(y_test, clf.predict_proba(X_test)[:, 1])
        print(f"  第 {i+1} 輪：train_auc={train_auc:.3f}, test_auc={test_auc:.3f}")

    # 檢查是否有有效的結果
    if len(results) == 0 or len(results_permuted) == 0:
        print("   無有效的訓練結果，跳過 Null Importance")
        return data, pd.DataFrame(columns=['變數', '原始gain', '反轉後gain'])

    imp_real_avg = pd.concat(results).groupby('feature')['gain'].mean().reset_index()
    imp_perm_avg = pd.concat(results_permuted).groupby('feature')['gain'].mean().reset_index()

    comparison = imp_real_avg.merge(imp_perm_avg, on='feature', suffixes=('_real', '_perm'))
    comparison['gain_real'] = comparison['gain_real'].fillna(0)
    comparison['gain_perm'] = comparison['gain_perm'].fillna(0)

    # 使用增益值打亂前後的絕對差值小於 50 作為條件
    comparison['gain_diff'] = (comparison['gain_real'] - comparison['gain_perm']).abs()
    noise_features = comparison[comparison['gain_diff'] < gain_threshold]['feature'].tolist()

    # 列出所有特徵的原始增益和打亂後增益
    dropped_info = pd.DataFrame({
        '變數': comparison['feature'].values,
        '原始gain': comparison['gain_real'].values,
        '反轉後gain': comparison['gain_perm'].values
    })
    # 添加狀態欄位，將剔除的特徵標記為 '去除'，保留的標記為 '保留'
    dropped_info['狀態'] = dropped_info.apply(
        lambda x: '去除' if np.abs(x['原始gain'] - x['反轉後gain']) < gain_threshold else '保留', axis=1
    )
    # 按原始增益降序排序
    dropped_info = dropped_info.sort_values('原始gain', ascending=False)
    dropped_info = dropped_info.reset_index(drop=True)
    # 添加原始增益排名欄位
    dropped_info['原始gain排名'] = range(1, len(dropped_info) + 1)

    data = data.drop(columns=[c for c in noise_features if c in data.columns])

    print(f"  已剔除 {len(noise_features)} 個雜訊特徵")
    return data, dropped_info


def _calc_single_psi(args):
    """計算單個機構和單個特徵的 PSI - NaN 作為獨立分箱"""
    org, train_month, test_month, train_n, test_n, f, data_ref, min_sample = args

    try:
        org_data = data_ref[data_ref['new_org'] == org]
        train_data = org_data[org_data['new_date_ym'] == train_month]
        test_data = org_data[org_data['new_date_ym'] == test_month]

        # 獲取資料
        train_vals = train_data[f].values
        test_vals = test_data[f].values

        # 標記 NaN
        train_nan_mask = pd.isna(train_vals)
        test_nan_mask = pd.isna(test_vals)

        # 用於分箱的非 NaN 值
        train_nonan = train_vals[~train_nan_mask]
        test_nonan = test_vals[~test_nan_mask]

        if len(train_nonan) < min_sample or len(test_nonan) < min_sample:
            return {
                '機構': org, '日期': f"{train_month}->{test_month}",
                '變數': f, 'PSI': None, '有效計算': 0,
                '樣本數': train_n
            }

        # 基於非 NaN 資料進行分箱 (10 箱)
        try:
            bins = pd.qcut(train_nonan, q=10, duplicates='drop', retbins=True)[1]
        except Exception:
            bins = pd.cut(train_nonan, bins=10, retbins=True)[1]

        # 計算各分箱占比 (包含 NaN 箱)
        train_counts = []
        test_counts = []

        for i in range(len(bins)):
            if i == 0:
                train_counts.append((~train_nan_mask & (train_vals <= bins[i])).sum())
                test_counts.append((~test_nan_mask & (test_vals <= bins[i])).sum())
            else:
                train_counts.append((~train_nan_mask & (train_vals > bins[i-1]) & (train_vals <= bins[i])).sum())
                test_counts.append((~test_nan_mask & (test_vals > bins[i-1]) & (test_vals <= bins[i])).sum())

        # NaN 箱
        train_counts.append(train_nan_mask.sum())
        test_counts.append(test_nan_mask.sum())

        # 轉換為占比
        train_pct = np.array(train_counts) / len(train_vals)
        test_pct = np.array(test_counts) / len(test_vals)

        # 避免 0 值
        train_pct = np.where(train_pct == 0, 1e-6, train_pct)
        test_pct = np.where(test_pct == 0, 1e-6, test_pct)

        # 計算 PSI
        psi = np.sum((test_pct - train_pct) * np.log(test_pct / train_pct))

        return {
            '機構': org, '日期': f"{train_month}->{test_month}",
            '變數': f, 'PSI': round(psi, 4), '有效計算': 1,
            '樣本數': train_n
        }
    except Exception as e:
        return {
            '機構': org, '日期': f"{train_month}->{test_month}",
            '變數': f, 'PSI': None, '有效計算': 0,
            '樣本數': train_n
        }


def drop_highpsi_features(data: pd.DataFrame, features: List[str],
                         psi_threshold: float = 0.1, max_months_ratio: float = 1/3,
                         max_orgs: int = 4, min_sample_per_month: int = 100, n_jobs: int = 4) -> tuple:
    """剔除高 PSI 特徵 - 按機構 + 逐月版本

    特徵層級多程序，機構層級迴圈，機構內特徵並行計算

    參數:
        psi_threshold: PSI 閾值，高於此值則視為不穩定
        max_months_ratio: 最大容忍月份比例，若超過此比例的月份 PSI 高於閾值，則記錄在處理表中
        max_orgs: 最大容忍機構數量，若超過此數量的機構不穩定，則記錄在處理表中
        min_sample_per_month: 每月最小樣本量

    返回:
        data: 剔除後的資料
        psi_detail: PSI 明細 (每個特徵在每個機構每個月份的 PSI 值)
        psi_process: PSI 處理表 (不符合條件的特徵)
    """
    orgs = data['new_org'].unique()

    # 建立任務列表：每個機構、每對月份、每個特徵
    tasks = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())

        if len(months) < 2:
            continue

        for i in range(len(months) - 1):
            train_month = months[i]
            test_month = months[i + 1]

            train_data = org_data[org_data['new_date_ym'] == train_month]
            test_data = org_data[org_data['new_date_ym'] == test_month]

            train_n = len(train_data)
            test_n = len(test_data)

            for f in features:
                tasks.append((org, train_month, test_month, train_n, test_n, f, data, min_sample_per_month))

    # 多程序計算 PSI (特徵層級並行)
    print(f"   PSI 計算：{len(tasks)} 個任務，使用 {n_jobs} 個程序")
    results = Parallel(n_jobs=n_jobs, verbose=0)(delayed(_calc_single_psi)(task) for task in tasks)

    psi_df = pd.DataFrame(results)

    if len(psi_df) == 0:
        return data, pd.DataFrame(columns=['變數', '機構', '月份', 'PSI值']), pd.DataFrame(columns=['變數', '處理原因'])

    # 篩選有效計算記錄
    valid_psi = psi_df[psi_df['有效計算'] == 1].copy()

    if len(valid_psi) == 0:
        return data, pd.DataFrame(columns=['變數', '機構', '月份', 'PSI值']), pd.DataFrame(columns=['變數', '處理原因'])

    # PSI 明細表：每個特徵在每個機構每個月份的 PSI 值
    # 將日期更改為單個月份，初始月 PSI 值為 0
    psi_detail = valid_psi[['機構', '日期', '變數', 'PSI']].copy()

    # 解析日期，提取測試月
    psi_detail['月份'] = psi_detail['日期'].apply(lambda x: x.split('->')[1] if '->' in x else x)
    psi_detail = psi_detail.rename(columns={'PSI': 'PSI值'})

    # 按特徵、機構、月份升序排序
    psi_detail = psi_detail.sort_values(['變數', '機構', '月份'], ascending=[True, True, True])

    # 獲取所有機構和月份
    all_orgs = sorted(psi_detail['機構'].unique())
    all_vars = sorted(psi_detail['變數'].unique())

    # 構建完整的 PSI 明細表 (包含初始月，PSI 值為 0)
    psi_detail_list = []
    for org in all_orgs:
        org_data = psi_detail[psi_detail['機構'] == org]
        if len(org_data) == 0:
            continue

        # 獲取該機構的所有月份
        months = sorted(org_data['月份'].unique())

        for var in all_vars:
            var_data = org_data[org_data['變數'] == var]
            if len(var_data) == 0:
                continue

            # 初始月 PSI 值為 0
            psi_detail_list.append({
                '機構': org,
                '變數': var,
                '月份': months[0],
                'PSI值': 0.0
            })

            # 後續月份 PSI 值為計算結果
            for i in range(1, len(months)):
                month = months[i]
                var_month_data = var_data[var_data['月份'] == month]
                if len(var_month_data) > 0:
                    psi_value = var_month_data['PSI值'].values[0]
                else:
                    psi_value = 0.0
                psi_detail_list.append({
                    '機構': org,
                    '變數': var,
                    '月份': month,
                    'PSI值': psi_value
                })

    psi_detail = pd.DataFrame(psi_detail_list)
    psi_detail = psi_detail[['機構', '變數', '月份', 'PSI值']]
    psi_detail = psi_detail.reset_index(drop=True)
    # 按特徵、機構、月份升序排序
    psi_detail = psi_detail.sort_values(['變數', '機構', '月份'], ascending=[True, True, True])
    psi_detail = psi_detail.reset_index(drop=True)

    # 標記不穩定
    valid_psi['不穩定'] = (valid_psi['PSI'] > psi_threshold).astype(int)

    # 彙總：每個特徵在每個機構不穩定月份數和總月份數
    org_summary = valid_psi.groupby(['機構', '變數']).agg(
        不穩定月份數=('不穩定', 'sum'),
        總月份數=('變數', 'count')
    ).reset_index()

    # 標記每個特徵在每個機構是否不穩定
    # 確保閾值至少為 1，避免機構月份較少時過於嚴格
    org_summary['不穩定閾值'] = org_summary['總月份數'].apply(
        lambda x: max(1, int(x * max_months_ratio))
    )
    org_summary['是否不穩定'] = org_summary['不穩定月份數'] >= org_summary['不穩定閾值']

    # 機構層級彙總：不穩定機構數量
    org_count = len(orgs)
    channel_summary = org_summary.groupby('變數').apply(
        lambda x: pd.Series({
            '機構數': org_count,
            '不穩定機構數': x['是否不穩定'].sum()
        })
    ).reset_index()

    # 標記需要處理的特徵
    channel_summary['需處理'] = channel_summary['不穩定機構數'] >= max_orgs
    channel_summary['處理原因'] = channel_summary.apply(
        lambda x: f'在 {x["不穩定機構數"]} 個機構中 PSI 不穩定' if x['需處理'] else '', axis=1
    )

    # 獲取每個特徵的不穩定機構清單
    unstable_orgs_dict = {}
    for var in org_summary['變數'].unique():
        var_orgs = org_summary[(org_summary['變數'] == var) & (org_summary['是否不穩定'] == True)]['機構'].tolist()
        unstable_orgs_dict[var] = var_orgs

    # PSI 處理表：不符合條件的特徵
    psi_process = channel_summary[channel_summary['需處理']].copy()
    psi_process['不穩定機構'] = psi_process['變數'].apply(lambda x: ','.join(unstable_orgs_dict.get(x, [])))
    psi_process = psi_process[['變數', '處理原因', '不穩定機構']]
    psi_process = psi_process.reset_index(drop=True)

    # 篩選待剔除特徵
    if len(psi_process) > 0 and '變數' in psi_process.columns:
        to_drop_vars = psi_process['變數'].tolist()
        data = data.drop(columns=[c for c in to_drop_vars if c in data.columns])

    return data, psi_detail, psi_process


def iv_distribution_by_org(iv_detail: pd.DataFrame, oos_orgs: list = None, iv_bins: list = [0, 0.02, 0.05, 0.1, float('inf')]) -> pd.DataFrame:
    """統計每個機構在不同 IV 區間的特徵數量和占比

    參數:
        iv_detail: IV 明細表 (包含特徵、整體、機構各列)
        oos_orgs: 貸外機構清單
        iv_bins: IV 區間邊界 [0, 0.02, 0.05, 0.1, inf]

    返回:
        IV 分佈統計表
    """
    if oos_orgs is None:
        oos_orgs = []

    # 獲取機構列 (排除 '變數' 和 '整體' 列)
    org_cols = [c for c in iv_detail.columns if c not in ['變數', '整體']]

    # 定義區間標籤
    bin_labels = ['[0, 0.02)', '[0.02, 0.05)', '[0.05, 0.1)', '[0.1, +∞)']

    result = []

    # 統計每個機構 (不含整體)
    for org in org_cols:
        org_iv = iv_detail[org].dropna()
        total_vars = len(org_iv)

        # 判斷機構類型
        org_type = '貸外' if org in oos_orgs else '建模'

        for i in range(len(iv_bins) - 1):
            lower = iv_bins[i]
            upper = iv_bins[i + 1]
            if upper == float('inf'):
                count = ((org_iv >= lower)).sum()
            else:
                count = ((org_iv >= lower) & (org_iv < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '機構': org,
                '類型': org_type,
                'IV區間': bin_labels[i],
                '變數個數': count,
                '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def psi_distribution_by_org(psi_detail: pd.DataFrame, oos_orgs: list = None, psi_bins: list = [0, 0.05, 0.1, float('inf')]) -> pd.DataFrame:
    """統計每個機構在不同 PSI 區間的特徵數量和占比

    參數:
        psi_detail: PSI 明細表 (包含機構、變數、月份、PSI值各列)
        oos_orgs: 貸外機構清單
        psi_bins: PSI 區間邊界 [0, 0.05, 0.1, inf]

    返回:
        PSI 分佈統計表
    """
    if oos_orgs is None:
        oos_orgs = []

    # 定義區間標籤
    bin_labels = ['[0, 0.05)', '[0.05, 0.1)', '[0.1, +∞)']

    result = []

    # 獲取所有機構
    orgs = psi_detail['機構'].unique()

    for org in orgs:
        org_data = psi_detail[psi_detail['機構'] == org]

        # 判斷機構類型
        org_type = '貸外' if org in oos_orgs else '建模'

        # 對於每個特徵，取其最大 PSI 值
        var_max_psi = org_data.groupby('變數')['PSI值'].max()
        total_vars = len(var_max_psi)

        for i in range(len(psi_bins) - 1):
            lower = psi_bins[i]
            upper = psi_bins[i + 1]
            if upper == float('inf'):
                count = ((var_max_psi >= lower)).sum()
            else:
                count = ((var_max_psi >= lower) & (var_max_psi < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '機構': org,
                '類型': org_type,
                'PSI區間': bin_labels[i],
                '變數個數': count,
                '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def value_ratio_distribution_by_org(data: pd.DataFrame, features: List[str],
                                     oos_orgs: list = None,
                                     value_bins: list = [0, 0.15, 0.35, 0.65, 0.95, 1.0]) -> pd.DataFrame:
    """統計每個機構在不同有值率區間的特徵數量和占比

    參數:
        data: 資料 (包含 new_org 列)
        features: 特徵列表
        oos_orgs: 貸外機構清單
        value_bins: 有值率區間邊界 [0, 0.15, 0.35, 0.65, 0.95, 1.0]

    返回:
        有值率分佈統計表
    """
    if oos_orgs is None:
        oos_orgs = []

    # 定義區間標籤
    bin_labels = ['[0, 15%)', '[15%, 35%)', '[35%, 65%)', '[65%, 95%)', '[95%, 100%]']

    result = []

    # 獲取所有機構
    orgs = data['new_org'].unique()

    for org in orgs:
        org_data = data[data['new_org'] == org]

        # 判斷機構類型
        org_type = '貸外' if org in oos_orgs else '建模'

        # 計算每個特徵的有值率 (非 NaN 占比)
        value_ratios = {}
        for f in features:
            if f in org_data.columns:
                non_null_count = org_data[f].notna().sum()
                total_count = len(org_data)
                value_ratios[f] = non_null_count / total_count if total_count > 0 else 0

        # 統計各區間特徵數量
        total_vars = len(value_ratios)
        for i in range(len(value_bins) - 1):
            lower = value_bins[i]
            upper = value_bins[i + 1]
            if upper == 1.0:
                count = sum(1 for v in value_ratios.values() if lower <= v <= upper)
            else:
                count = sum(1 for v in value_ratios.values() if lower <= v < upper)
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '機構': org,
                '類型': org_type,
                '有值率區間': bin_labels[i],
                '變數個數': count,
                '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def calculate_iv_by_org(data: pd.DataFrame, features: List[str],
                        n_jobs: int = 4) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """按機構和整體計算 IV

    返回:
        iv_by_org: 分機構 IV 明細
        iv_overall: 整體 IV
    """
    from references.func import calculate_iv

    orgs = data['new_org'].unique()

    # 整體 IV
    iv_overall = calculate_iv(data, features, n_jobs=n_jobs)
    iv_overall['類型'] = '整體'

    # 分機構 IV
    iv_by_org = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, n_jobs=1)  # 單機構單程序計算
        if len(org_iv) > 0:  # 僅添加非空結果
            org_iv['機構'] = org
            org_iv['類型'] = '分機構'
            iv_by_org.append(org_iv)

    iv_by_org = pd.concat(iv_by_org, ignore_index=True) if iv_by_org else pd.DataFrame(columns=['變數', 'IV', '機構', '類型'])

    return iv_by_org, iv_overall


def calculate_psi_detail(data: pd.DataFrame, features: List[str],
                         max_psi: float = 0.1, min_months_unstable: int = 3,
                         min_sample: int = 100, n_jobs: int = 4) -> tuple:
    """計算每個特徵在每個機構的逐月 PSI 明細，並標記是否剔除

    返回:
        data: 剔除後的資料
        dropped: 剔除變數彙總
        psi_summary: 完整的 PSI 明細 (含剔除標記)
    """
    orgs = data['new_org'].unique()

    # 建立任務
    tasks = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())

        if len(months) < 2:
            continue

        for i in range(len(months) - 1):
            train_month = months[i]
            test_month = months[i + 1]

            train_data = org_data[org_data['new_date_ym'] == train_month]
            test_data = org_data[org_data['new_date_ym'] == test_month]

            train_n = len(train_data)
            test_n = len(test_data)

            for f in features:
                tasks.append((org, train_month, test_month, train_n, test_n, f, data, min_sample))

    # 多程序計算
    print(f"   PSI 計算：{len(tasks)} 個任務，使用 {n_jobs} 個程序")
    results = Parallel(n_jobs=n_jobs, verbose=0)(delayed(_calc_single_psi)(task) for task in tasks)

    psi_df = pd.DataFrame(results)

    if len(psi_df) == 0:
        return data, pd.DataFrame(columns=['變數', '機構數', '不穩定機構數', '原因']), pd.DataFrame(columns=['變數', '機構數', '不穩定機構數', '是否剔除', '去除條件'])

    # 篩選有效計算記錄
    valid_psi = psi_df[psi_df['有效計算'] == 1].copy()

    if len(valid_psi) == 0:
        return data, pd.DataFrame(columns=['變數', '機構數', '不穩定機構數', '原因']), pd.DataFrame(columns=['變數', '機構數', '不穩定機構數', '是否剔除', '去除條件'])

    # 標記不穩定
    valid_psi['不穩定'] = (valid_psi['PSI'] > max_psi).astype(int)

    # 彙總：每個特徵在每個機構不穩定月份數
    org_summary = valid_psi.groupby(['機構', '變數'])['不穩定'].sum().reset_index()
    org_summary.columns = ['機構', '變數', '不穩定月份數']

    # 機構層級彙總：不穩定月份數大於等於 min_months_unstable 的特徵
    org_count = len(orgs)
    channel_summary = org_summary.groupby('變數').apply(
        lambda x: pd.Series({
            '機構數': org_count,
            '不穩定機構數': (x['不穩定月份數'] >= min_months_unstable).sum()
        })
    ).reset_index()

    # 標記需要剔除的特徵 (超過 1/3 機構不穩定)
    channel_summary['需剔除'] = channel_summary['不穩定機構數'] > (channel_summary['機構數'] / 3)
    channel_summary['是否剔除'] = channel_summary['需剔除'].astype(int)
    channel_summary['去除條件'] = channel_summary.apply(
        lambda x: f'超過 1/3 的 {org_count} 個機構在連續 {min_months_unstable} 個月 PSI > {max_psi}' if x['需剔除'] else '', axis=1
    )

    # 篩選待剔除特徵
    if len(channel_summary) > 0 and '變數' in channel_summary.columns:
        to_drop_vars = channel_summary[channel_summary['需剔除']]['變數'].tolist()
        data = data.drop(columns=[c for c in to_drop_vars if c in data.columns])

    # 整理剔除資訊 (僅返回剔除變數)
    dropped = channel_summary[channel_summary['需剔除']].copy()
    dropped['原因'] = f'超過 1/3 的 {org_count} 個機構在連續 {min_months_unstable} 個月 PSI > {max_psi}'

    return data, dropped[['變數', '機構數', '不穩定機構數', '原因']], channel_summary[['變數', '機構數', '不穩定機構數', '是否剔除', '去除條件']]


def export_cleaning_report(filepath: str, steps: list,
                           iv_detail: pd.DataFrame = None,
                           iv_process: pd.DataFrame = None,
                           psi_detail: pd.DataFrame = None,
                           psi_process: pd.DataFrame = None,
                           params: dict = None,
                           iv_distribution: pd.DataFrame = None,
                           psi_distribution: pd.DataFrame = None,
                           value_ratio_distribution: pd.DataFrame = None):
    """匯出清洗報告至 xlsx - 每個步驟一個工作表

    參數:
        filepath: 輸出路徑
        steps: 清洗步驟列表 [(步驟名稱, DataFrame), ...]
        iv_detail: IV 明細 (每個特徵在每個機構及整體的 IV 值)
        iv_process: IV 處理表 (不符合條件的特徵)
        psi_detail: PSI 明細 (每個特徵在每個機構每個月份的 PSI 值)
        psi_process: PSI 處理表 (不符合條件的特徵)
        params: 超參數字典，用於動態生成條件
        iv_distribution: IV 分佈統計表
        psi_distribution: PSI 分佈統計表
        value_ratio_distribution: 有值率分佈統計表
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filepath)
    except Exception:
        wb = Workbook()
        wb.remove(wb.active)

    # 彙總工作表 - 僅顯示真實的過濾步驟
    if '彙總' in wb.sheetnames:
        del wb['彙總']
    ws = wb.create_sheet('彙總', 0)
    ws['A1'] = '資料清洗報告'
    ws['A2'] = f'生成時間：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A4'] = '步驟'
    ws['B4'] = '操作內容'
    ws['C4'] = '操作結果'
    ws['D4'] = '條件'

    # 僅顯示真實的過濾步驟 (排除明細和分佈統計)
    filter_steps = [
        'Step4-異常月份處理', 'Step6-高缺失率處理', 'Step7-IV處理',
        'Step8-PSI處理', 'Step9-null importance處理', 'Step10-高相關性剔除'
    ]

    # 需要排除的步驟 (明細和分佈統計)
    exclude_steps = [
        'Step7-IV明細', 'Step7-IV分佈統計', 'Step8-PSI明細',
        'Step8-PSI分佈統計', 'Step5-有值率分佈統計'
    ]

    # 需要顯示剔除數量的步驟
    show_drop_count_steps = ['分離OOS資料']

    # 僅顯示參數標準的步驟 (無操作結果)
    show_param_only_steps = ['機構樣本統計', '缺失率明細']

    # 添加備註：每個步驟獨立執行
    ws['A3'] = '註：每個過濾步驟皆為獨立執行，不會刪除資料，僅記錄不符合條件的特徵統計'

    # 獲取參數，未提供則使用預設值
    if params is None:
        params = {}

    min_ym_bad_sample = params.get('min_ym_bad_sample', 10)
    min_ym_sample = params.get('min_ym_sample', 500)
    missing_ratio = params.get('missing_ratio', 0.6)
    overall_iv_threshold = params.get('overall_iv_threshold', 0.1)
    org_iv_threshold = params.get('org_iv_threshold', 0.1)
    max_org_threshold = params.get('max_org_threshold', 2)
    psi_threshold = params.get('psi_threshold', 0.1)
    max_months_ratio = params.get('max_months_ratio', 1/3)
    max_orgs = params.get('max_orgs', 4)
    gain_threshold = params.get('gain_threshold', 50)

    step_num = 1
    for name, df in steps:
        # 跳過明細和分佈統計步驟
        if name in exclude_steps:
            continue

        # 移除操作內容中的 StepX- 前綴
        display_name = name.replace('Step4-', '').replace('Step6-', '').replace('Step7-', '').replace('Step8-', '').replace('Step9-', '').replace('Step10-', '')

        # 僅顯示參數標準步驟 (無操作結果)
        if name in show_param_only_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            result = ''
            # 條件：顯示參數標準
            if name == '機構樣本統計':
                condition = '每個機構的樣本數及壞樣本率統計'
            elif name == '缺失率明細':
                condition = '計算每個特徵的缺失率'
            else:
                condition = ''
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
        # 顯示需要顯示剔除數量的步驟
        elif name in show_drop_count_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            if df is not None and len(df) > 0:
                if name == '分離OOS資料':
                    # 特殊處理：顯示 OOS 和建模樣本數量
                    if '變數' in df.columns and '數量' in df.columns:

                        oos_count = df[df['變數'] == 'OOS樣本']['數量'].values[0] if len(df[df['變數'] == 'OOS樣本']) > 0 else 0
                        model_count = df[df['變數'] == '建模樣本']['數量'].values[0] if len(df[df['變數'] == '建模樣本']) > 0 else 0
                        result = f'OOS 樣本 {oos_count}, 建模樣本 {model_count}'
                    else:
                        result = f'{len(df)} 行'
                elif '變數' in df.columns:
                    result = f'已剔除 {len(df)} 個特徵'
                else:
                    result = f'已剔除 {len(df)}'
                condition = ''
            else:
                result = '空'
                condition = ''
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
        elif name in filter_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)

            # 生成操作結果和條件
            if df is not None and len(df) > 0:
                if name == 'Step4-異常月份處理':
                    # 操作結果：剔除的月份
                    if '年月' in df.columns:
                        result = '已剔除 ' + ','.join(df['年月'].astype(str).tolist())
                    else:
                        result = '已剔除 ' + ','.join(df.iloc[:, 0].astype(str).tolist())
                    # 條件：參數標準
                    condition = f'壞樣本數小於 {min_ym_bad_sample} 或總樣本數小於 {min_ym_sample} 的月份將被剔除 (獨立執行)'
                elif name == 'Step6-高缺失率處理':
                    # 操作結果：剔除特徵數量
                    if '變數' in df.columns:
                        result = f'已剔除 {len(df)} 個特徵'
                    else:
                        result = f'已剔除 {len(df)}'
                    # 條件：參數標準
                    condition = f'整體缺失率大於 {missing_ratio} 的特徵將被剔除 (獨立執行)'
                elif name == 'Step7-IV處理':
                    # 操作結果：剔除特徵數量
                    if '變數' in df.columns:
                        result = f'已剔除 {len(df)} 個特徵'
                    else:
                        result = f'已剔除 {len(df)}'
                    # 條件：參數標準
                    condition = f'整體 IV 小於 {overall_iv_threshold} 或在 {max_org_threshold} 個或更多機構中 IV 小於 {org_iv_threshold} 的特徵將被剔除 (獨立執行)'
                elif name == 'Step8-PSI處理':
                    # 操作結果：剔除特徵數量
                    if '變數' in df.columns:
                        result = f'已剔除 {len(df)} 個特徵'
                    else:
                        result = f'已剔除 {len(df)}'
                    # 條件：參數標準
                    condition = f'PSI 閾值 {psi_threshold}，若機構超過 {max_months_ratio:.0%} 的月份 PSI 大於 {psi_threshold} 則視為機構不穩定，若不穩定機構超過 {max_orgs} 個則剔除特徵 (獨立執行)'
                elif name == 'Step9-null importance處理':
                    # 操作結果：剔除特徵數量
                    if '變數' in df.columns:
                        result = f'已剔除 {len(df[df["狀態"]=="去除"])} 個特徵'
                    else:
                        result = f'已剔除 {len(df)}'
                    # 條件：參數標準
                    condition = f'增益值在打亂前後的絕對差值小於 {gain_threshold} 的特徵將被識別為雜訊並剔除 (獨立執行)'
                elif name == 'Step10-高相關性剔除':
                    # 操作結果：剔除特徵數量
                    if '變數' in df.columns:
                        result = f'已剔除 {len(df)} 個特徵'
                    else:
                        result = f'已剔除 {len(df)}'
                    # 條件：參數標準
                    max_corr = params.get('max_corr', 0.9)
                    top_n_keep = params.get('top_n_keep', 20)
                    condition = f'相關性大於 {max_corr} 的特徵將被剔除，按原始 gain 排名保留前 {top_n_keep} 個特徵 (獨立執行)'
                else:
                    result = '已剔除 ' + str(len(df))
                    condition = ''
            else:
                result = '空'
                condition = ''

            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1

    # 計算總剔除特徵數 (取各步驟剔除特徵的聯集)
    all_dropped_vars = set()
    for name, df in steps:
        if name in filter_steps and df is not None and len(df) > 0 and '變數' in df.columns:
            if name == 'Step9-null importance處理':
                # null importance 處理需過濾狀態為 "去除" 的特徵
                dropped_vars = df[df['狀態'] == '去除']['變數'].tolist()
            else:
                dropped_vars = df['變數'].tolist()
            # 取聯集 (去重)
            all_dropped_vars = all_dropped_vars.union(set(dropped_vars))

    # 添加最後一行統計
    final_step_num = step_num
    ws.cell(4+final_step_num, 1, final_step_num)
    ws.cell(4+final_step_num, 2, '最終剔除特徵統計')
    ws.cell(4+final_step_num, 3, f'累計剔除 {len(all_dropped_vars)} 個特徵 (各步驟聯集)')
    ws.cell(4+final_step_num, 4, '每個步驟獨立執行，最終剔除特徵為各步驟剔除特徵的聯集')

    # 各步驟明細 (按步驟順序建立工作表)
    # 定義工作表建立順序
    sheet_order = [
        '機構樣本統計', '分離OOS資料', 'Step4-異常月份處理', '缺失率明細',
        'Step5-有值率分佈統計', 'Step6-高缺失率處理', 'Step7-IV明細', 'Step7-IV處理',
        'Step7-IV分佈統計', 'Step8-PSI明細', 'Step8-PSI處理', 'Step8-PSI分佈統計',
        'Step9-null importance處理', 'Step10-高相關性剔除'
    ]

    # 按順序建立工作表
    for sheet_name in sheet_order:
        # 從 steps 中尋找對應的 DataFrame
        df = None
        for name, step_df in steps:
            if name == sheet_name:
                df = step_df
                break

        if df is not None:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws_detail = wb.create_sheet(sheet_name)

            for j, col in enumerate(df.columns):
                ws_detail.cell(1, j+1, col)

            for i, row in df.iterrows():
                for j, val in enumerate(row):
                    # 直接寫入值，避免字元轉義問題
                    ws_detail.cell(i+2, j+1, val if val is not None else '')

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font

    wb.save(filepath)
    print(f"報告已儲存：{filepath}")
